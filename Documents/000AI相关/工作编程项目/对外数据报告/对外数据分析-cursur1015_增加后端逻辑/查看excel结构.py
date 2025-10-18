#!/usr/bin/env python3
# 简单查看Excel文件结构的脚本

import sys
import os

try:
    # 尝试导入openpyxl
    import openpyxl
    from openpyxl import load_workbook
    
    excel_file = '2025109移山科技循环10次采集(20251014)对外报表_副本.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"文件不存在: {excel_file}")
        sys.exit(1)
    
    # 加载工作簿
    wb = load_workbook(excel_file, read_only=True)
    
    print("Excel文件包含的sheet:")
    for i, sheet_name in enumerate(wb.sheetnames):
        print(f"{i+1}. {sheet_name}")
    
    print("\n" + "="*50)
    
    # 查看每个sheet的前几行
    for sheet_name in wb.sheetnames[:5]:
        print(f"\n【{sheet_name}】数据预览:")
        ws = wb[sheet_name]
        
        # 获取数据范围
        max_row = min(5, ws.max_row) if ws.max_row else 0
        max_col = min(10, ws.max_column) if ws.max_column else 0
        
        print(f"数据范围: {ws.max_row} rows x {ws.max_column} columns")
        
        if max_row > 0 and max_col > 0:
            # 打印表头
            headers = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value is not None else "")
            print("列名:", headers)
            
            # 打印前几行数据
            for row in range(2, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                print(f"第{row-1}行:", row_data)
        else:
            print("该sheet为空或无数据")
        
        print("-" * 30)
    
    wb.close()

except ImportError:
    print("需要安装openpyxl库")
    print("请运行: pip install openpyxl")
except Exception as e:
    print(f"读取Excel文件时出错: {e}")